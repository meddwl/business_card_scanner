import os
import json
import shutil
import time
from pathlib import Path
from datetime import datetime

from google import genai
from google.genai import types
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

# ── Config ────────────────────────────────────────────────────────────────────

load_dotenv()

INPUT_DIR     = Path("input")
PROCESSED_DIR = Path("processed")
EXCEL_FILE    = Path(os.getenv("EXCEL_FILE", "contacts.xlsx"))

SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".heic", ".heif"}

HEADERS = ["Full Name", "Position", "Company", "Phone", "Email", "Processed At", "Source File"]

# ── Gemini setup ──────────────────────────────────────────────────────────────

SYSTEM_INSTRUCTION = """You are a business card data extraction specialist.

Your ONLY job is to read a business card image and return a single JSON object.

Rules:
- Return ONLY raw JSON. No markdown, no code fences, no explanation.
- Always include exactly these five keys: full_name, position, company, phone, email.
- If a field cannot be found on the card, use an empty string "".
- For phone, prefer mobile over landline. Include country code if visible.
- Do not invent or guess values that are not clearly on the card.

Output format:
{"full_name": "...", "position": "...", "company": "...", "phone": "...", "email": "..."}
"""

def get_gemini_client():
    """Create and return a Gemini client using the new google.genai SDK."""
    return genai.Client(api_key=os.getenv("GEMINI_API_KEY"))

# ── Excel ─────────────────────────────────────────────────────────────────────

def get_workbook_and_sheet():
    """Open existing Excel file or create a new one with a header row."""
    if EXCEL_FILE.exists():
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        print(f"  📂 Opened existing file: {EXCEL_FILE}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Contacts"
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)
        print(f"  📄 Created new file: {EXCEL_FILE}")
    return wb, ws


def append_contact(wb, ws, contact: dict, filename: str):
    """Append one contact row and save the file."""
    row = [
        contact.get("full_name", ""),
        contact.get("position", ""),
        contact.get("company", ""),
        contact.get("phone", ""),
        contact.get("email", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        filename,
    ]
    ws.append(row)
    wb.save(EXCEL_FILE)

# ── Gemini Vision ─────────────────────────────────────────────────────────────

def load_image_part(image_path: Path) -> types.Part:
    """Return a Gemini-compatible image Part from a file."""
    mime_map = {
        ".jpg":  "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png":  "image/png",
        ".webp": "image/webp",
        ".gif":  "image/gif",
        ".heic": "image/heic",
        ".heif": "image/heif",
    }
    mime_type = mime_map.get(image_path.suffix.lower(), "image/jpeg")
    with open(image_path, "rb") as f:
        raw_bytes = f.read()
    return types.Part.from_bytes(data=raw_bytes, mime_type=mime_type)


def extract_contact(client, image_path: Path) -> dict | None:
    """Send image to Gemini and return extracted contact as a dict.
    Retries up to 3 times on rate limit errors (429).
    """
    image_part = load_image_part(image_path)

    for attempt in range(3):
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=["Extract all contact information from this business card.", image_part],
                config=types.GenerateContentConfig(
                    system_instruction=SYSTEM_INSTRUCTION,
                    response_mime_type="application/json",
                    temperature=0,
                    max_output_tokens=512,
                ),
            )

            raw = response.text.strip()

            # Strip accidental markdown fences (safety net)
            if raw.startswith("```"):
                raw = raw.split("```")[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                raw = raw.strip()

            return json.loads(raw)

        except json.JSONDecodeError as e:
            print(f"  ⚠️  Could not parse JSON from Gemini's response: {e}")
            print(f"  Raw response: {raw}")
            return None

        except Exception as e:
            if "429" in str(e) and attempt < 2:
                wait = 60 * (attempt + 1)  # 60s on first retry, 120s on second
                print(f"  ⏳ Rate limit hit — waiting {wait}s before retry {attempt + 2}/3...")
                time.sleep(wait)
            else:
                print(f"  ❌ API error: {e}")
                return None

# ── Main Loop ─────────────────────────────────────────────────────────────────

def main():
    INPUT_DIR.mkdir(exist_ok=True)
    PROCESSED_DIR.mkdir(exist_ok=True)

    images = [f for f in INPUT_DIR.iterdir() if f.suffix.lower() in SUPPORTED_EXTENSIONS]

    if not images:
        print("📭 No images found in the 'input/' folder. Drop some business card photos there and run again.")
        return

    print(f"🔍 Found {len(images)} image(s) to process.\n")

    print("🤖 Initialising Gemini 2.5 Flas ...")
    try:
        client = get_gemini_client()
        print("  ✅ Client ready.\n")
    except Exception as e:
        print(f"  ❌ Could not initialise Gemini: {e}")
        return

    print("📊 Opening Excel file...")
    try:
        wb, ws = get_workbook_and_sheet()
        print("  ✅ Ready.\n")
    except Exception as e:
        print(f"  ❌ Could not open/create Excel file: {e}")
        return

    success_count = 0
    fail_count = 0

    for image_path in sorted(images):
        print(f"📇 Processing: {image_path.name}")

        contact = extract_contact(client, image_path)

        if contact:
            print(f"  👤 {contact.get('full_name') or '(no name)'} — {contact.get('position') or '(no position)'} @ {contact.get('company') or '(no company)'}")
            print(f"  📞 {contact.get('phone') or '—'}   ✉️  {contact.get('email') or '—'}")

            try:
                append_contact(wb, ws, contact, image_path.name)
                print(f"  💾 Saved to {EXCEL_FILE}.")
            except Exception as e:
                print(f"  ❌ Failed to write to Excel: {e}")
                fail_count += 1
                continue

            dest = PROCESSED_DIR / image_path.name
            if dest.exists():
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = PROCESSED_DIR / f"{image_path.stem}_{timestamp}{image_path.suffix}"

            shutil.move(str(image_path), dest)
            print(f"  📁 Moved to processed/ → {dest.name}\n")
            success_count += 1
        else:
            print("  ❌ Extraction failed — image skipped (stays in input/).\n")
            fail_count += 1

    print("─" * 50)
    print(f"✅ Done!  {success_count} card(s) imported  |  {fail_count} failed")
    if success_count:
        print(f"📊 Open '{EXCEL_FILE}' to see your new contacts.")


if __name__ == "__main__":
    main()
